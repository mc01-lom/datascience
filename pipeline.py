"""
pipeline_rfi_to_painel.py
─────────────────────────────────────────────────────────────────────────────
Lê as abas de DRE de cada RFI de fornecedor e alimenta as abas de destino
do Painel com os dados mensais, usando mapeamento posicional declarado
inteiramente no config.xlsx.

Configuração : edite config.xlsx (abas Geral, Fornecedores, Colunas, Mapeamento)
Uso          : python pipeline_rfi_to_painel.py
─────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations
from pathlib import Path
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

CONFIG_FILE = Path(__file__).parent / "config.xlsx"


# ═══════════════════════════════════════════════════════════════════════════
# 1. ESTRUTURAS DE DADOS
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class Regra:
    """Uma linha da aba Mapeamento do config."""
    aba_origem  : str
    aba_destino : str
    ori_ini     : int   # linha origem início (1-indexed)
    ori_fim     : int   # linha origem fim    (1-indexed)
    dest_ini    : int   # linha destino início (1-indexed)
    dest_fim    : int   # linha destino fim    (1-indexed)
    acao        : str   # "Copiar" | "Somar"

    def __post_init__(self):
        if self.acao == "Copiar":
            tam_ori  = self.ori_fim  - self.ori_ini
            tam_dest = self.dest_fim - self.dest_ini
            if tam_ori != tam_dest:
                raise ValueError(
                    f"Regra Copiar inválida ({self.aba_origem} L{self.ori_ini}–{self.ori_fim} "
                    f"→ {self.aba_destino} L{self.dest_ini}–{self.dest_fim}): "
                    f"intervalos de tamanhos diferentes ({tam_ori+1} vs {tam_dest+1})."
                )
        elif self.acao == "Somar":
            if self.dest_ini != self.dest_fim:
                raise ValueError(
                    f"Regra Somar inválida ({self.aba_origem} L{self.ori_ini}–{self.ori_fim} "
                    f"→ {self.aba_destino} L{self.dest_ini}): "
                    f"LinhaDestIni deve ser igual a LinhaDestFim para Somar."
                )
        else:
            raise ValueError(f"Ação desconhecida: '{self.acao}'. Use 'Copiar' ou 'Somar'.")

    @property
    def pares(self) -> list[tuple[int, int, str]]:
        """
        Retorna lista de (linha_origem, linha_destino, acao) para cada linha do bloco.
        Para Somar, todas as origens apontam para dest_ini.
        """
        if self.acao == "Copiar":
            return [
                (self.ori_ini + i, self.dest_ini + i, "Copiar")
                for i in range(self.ori_fim - self.ori_ini + 1)
            ]
        else:
            return [
                (ori, self.dest_ini, "Somar")
                for ori in range(self.ori_ini, self.ori_fim + 1)
            ]


# ═══════════════════════════════════════════════════════════════════════════
# 2. LEITURA DO CONFIG
# ═══════════════════════════════════════════════════════════════════════════

def load_config() -> tuple[Path, Path, dict, dict[int, tuple[int, int]], list[Regra]]:
    """
    Retorna:
        painel_file : Path
        output_file : Path
        suppliers   : { nome: { "rfi_file": Path, "anos": { ano: (col_dest_ini, col_dest_fim) } } }
        col_origem  : { ano: (col_ori_ini, col_ori_fim) }   — 1-indexed
        regras      : [ Regra, ... ]
    """
    if not CONFIG_FILE.exists():
        raise FileNotFoundError(f"Config não encontrado: {CONFIG_FILE}")

    wb   = load_workbook(CONFIG_FILE, data_only=True)
    base = Path(__file__).parent

    # ── Aba Geral ────────────────────────────────────────────────────────
    geral = {}
    for row in wb["Geral"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and not str(row[0]).startswith("ℹ"):
            geral[str(row[0]).strip()] = str(row[1]).strip()

    painel_file = base / geral.get("Arquivo Painel", "Painel_BUPJ.xlsx")
    output_file = base / geral.get("Arquivo Saída",  "Painel_BUPJ_atualizado.xlsx")

    # ── Aba Fornecedores ─────────────────────────────────────────────────
    suppliers: dict[str, dict] = {}
    for row in wb["Fornecedores"].iter_rows(min_row=2, values_only=True):
        nome, rfi_file, ano, col_ini, col_fim = (tuple(row) + (None,) * 5)[:5]
        if not all([nome, rfi_file, ano, col_ini, col_fim]):
            continue
        if str(nome).startswith("ℹ"):
            continue
        nome = str(nome).strip()
        if nome not in suppliers:
            suppliers[nome] = {"rfi_file": base / str(rfi_file).strip(), "anos": {}}
        suppliers[nome]["anos"][int(ano)] = (
            column_index_from_string(str(col_ini).strip()),
            column_index_from_string(str(col_fim).strip()),
        )

    # ── Aba Colunas ──────────────────────────────────────────────────────
    col_origem: dict[int, tuple[int, int]] = {}
    for row in wb["Colunas"].iter_rows(min_row=3, values_only=True):
        ano, ini, fim = (tuple(row) + (None,) * 3)[:3]
        if not all([ano, ini, fim]):
            continue
        if str(ini).startswith("←") or str(fim).startswith("←"):
            print(f"  [AVISO] Colunas de origem para {ano} não preenchidas — ano ignorado.")
            continue
        col_origem[int(ano)] = (
            column_index_from_string(str(ini).strip()),
            column_index_from_string(str(fim).strip()),
        )

    # ── Aba Mapeamento ───────────────────────────────────────────────────
    regras: list[Regra] = []
    erros_regras = []
    for i, row in enumerate(wb["Mapeamento"].iter_rows(min_row=3, values_only=True), start=3):
        aba_ori, aba_dest, ori_ini, ori_fim, dest_ini, dest_fim, acao = (tuple(row) + (None,) * 7)[:7]
        if not all([aba_ori, aba_dest, ori_ini, ori_fim, dest_ini, dest_fim, acao]):
            continue
        try:
            regras.append(Regra(
                aba_origem  = str(aba_ori).strip(),
                aba_destino = str(aba_dest).strip(),
                ori_ini     = int(ori_ini),
                ori_fim     = int(ori_fim),
                dest_ini    = int(dest_ini),
                dest_fim    = int(dest_fim),
                acao        = str(acao).strip(),
            ))
        except ValueError as e:
            erros_regras.append(f"    Linha {i} da aba Mapeamento: {e}")

    wb.close()

    if erros_regras:
        print("\n[ERRO] Regras inválidas no config:")
        for e in erros_regras:
            print(e)
        raise SystemExit(1)

    return painel_file, output_file, suppliers, col_origem, regras


# ═══════════════════════════════════════════════════════════════════════════
# 3. LEITURA DO RFI (posicional)
# ═══════════════════════════════════════════════════════════════════════════

def read_rfi_aba(
    rfi_path  : Path,
    aba       : str,
    anos_ori  : dict[int, tuple[int, int]],   # { ano: (col_ini, col_fim) } — cols do RFI
    regras_aba: list[Regra],
) -> dict[str, dict[int, dict[int, list[float]]]]:
    """
    Lê uma aba do RFI e retorna os dados mapeados.

    Retorna:
        { aba_destino: { linha_destino: { ano: [v_jan..v_dez] } } }

    Para Somar, os valores de múltiplas linhas de origem são acumulados
    na mesma linha de destino.
    """
    try:
        wb = load_workbook(rfi_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Não foi possível abrir '{rfi_path.name}': {e}") from e

    if aba not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba '{aba}' não encontrada em '{rfi_path.name}'.")

    ws   = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # resultado[aba_destino][linha_destino][ano] = [vals]
    resultado: dict[str, dict[int, dict[int, list[float]]]] = {}

    for regra in regras_aba:
        for ori_row, dest_row, acao in regra.pares:
            row_idx = ori_row - 1  # 0-indexed
            if row_idx >= len(rows):
                print(f"      [AVISO] Linha {ori_row} não existe na aba '{aba}' — pulando.")
                continue

            row_data = rows[row_idx]

            for ano, (col_ini, col_fim) in anos_ori.items():
                vals = []
                for col_idx in range(col_ini - 1, col_fim):  # 0-indexed
                    raw = row_data[col_idx] if col_idx < len(row_data) else None
                    vals.append(float(raw) if isinstance(raw, (int, float)) else 0.0)

                dest = regra.aba_destino
                if dest not in resultado:
                    resultado[dest] = {}
                if dest_row not in resultado[dest]:
                    resultado[dest][dest_row] = {}

                if acao == "Copiar":
                    resultado[dest][dest_row][ano] = vals
                else:  # Somar
                    if ano not in resultado[dest][dest_row]:
                        resultado[dest][dest_row][ano] = [0.0] * len(vals)
                    for m, v in enumerate(vals):
                        resultado[dest][dest_row][ano][m] += v

    return resultado


# ═══════════════════════════════════════════════════════════════════════════
# 4. ESCRITA NO PAINEL
# ═══════════════════════════════════════════════════════════════════════════

def write_to_painel(
    wb_painel : object,
    aba_dest  : str,
    dados     : dict[int, dict[int, list[float]]],   # { linha_dest: { ano: [vals] } }
    anos_dest : dict[int, tuple[int, int]],           # { ano: (col_ini, col_fim) } no Painel
) -> int:
    """
    Grava dados em uma aba do Painel. Retorna número de linhas gravadas.
    """
    if aba_dest not in wb_painel.sheetnames:
        print(f"      [AVISO] Aba '{aba_dest}' não encontrada no Painel — pulando.")
        return 0

    ws = wb_painel[aba_dest]
    gravadas = 0

    for painel_row, anos_dict in dados.items():
        for ano, vals in anos_dict.items():
            if ano not in anos_dest:
                continue
            col_ini, col_fim = anos_dest[ano]
            for i, col in enumerate(range(col_ini, col_fim + 1)):
                if i < len(vals):
                    ws.cell(row=painel_row, column=col).value = vals[i]
        gravadas += 1

    return gravadas


# ═══════════════════════════════════════════════════════════════════════════
# 5. DESCOBERTA E CONFIRMAÇÃO DE RFIs
# ═══════════════════════════════════════════════════════════════════════════

def discover_rfis() -> list[Path]:
    return sorted(Path(__file__).parent.glob("RFI_*.xlsx"))


def confirm_rfis(rfis: list[Path], suppliers: dict) -> bool:
    mapped_names = {cfg["rfi_file"].name for cfg in suppliers.values()}

    print("=" * 60)
    print("  RFIs encontrados na pasta:")
    print("=" * 60)
    if not rfis:
        print("  [NENHUM arquivo RFI_*.xlsx encontrado]")
        return False

    for i, rfi in enumerate(rfis, 1):
        aviso = "" if rfi.name in mapped_names else "  ⚠ não mapeado no config"
        print(f"  {i}. {rfi.name}{aviso}")

    ausentes = [
        (nome, cfg["rfi_file"].name)
        for nome, cfg in suppliers.items()
        if not cfg["rfi_file"].exists()
    ]
    if ausentes:
        print()
        print("  [AVISO] Fornecedores no config sem RFI na pasta:")
        for nome, fname in ausentes:
            print(f"    • {nome} → {fname}")

    print("=" * 60)
    resposta = input("  Digite OK para processar ou qualquer outra tecla para cancelar: ").strip()
    return resposta.upper() == "OK"


# ═══════════════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    # ── Carregar config ──────────────────────────────────────────────────
    try:
        painel_file, output_file, suppliers, col_origem, regras = load_config()
    except (FileNotFoundError, SystemExit) as e:
        print(f"[ERRO] {e}")
        return

    anos_globais = sorted({ano for cfg in suppliers.values() for ano in cfg["anos"]})
    abas_origem  = sorted({r.aba_origem for r in regras})

    print(f"Configuração carregada:")
    print(f"  • {len(suppliers)} fornecedor(es) | anos: {anos_globais}")
    print(f"  • {len(regras)} regra(s) de mapeamento | abas DRE: {abas_origem}")
    print(f"  • Colunas de origem configuradas para: {sorted(col_origem)}")

    # ── Descoberta e confirmação ─────────────────────────────────────────
    rfis = discover_rfis()
    if not confirm_rfis(rfis, suppliers):
        print("\nOperação cancelada pelo usuário.")
        return

    # ── Agrupar regras por aba de origem ─────────────────────────────────
    regras_por_aba: dict[str, list[Regra]] = {}
    for r in regras:
        regras_por_aba.setdefault(r.aba_origem, []).append(r)

    # ── Carregar Painel ──────────────────────────────────────────────────
    print(f"\n[1/3] Carregando Painel base: {painel_file.name}...")
    if not painel_file.exists():
        print(f"[ERRO] Painel não encontrado: {painel_file}")
        return
    wb_painel = load_workbook(painel_file)
    print(f"      Painel carregado. Abas: {wb_painel.sheetnames}")

    # ── Processar fornecedores ───────────────────────────────────────────
    total       = sum(1 for cfg in suppliers.values() if cfg["rfi_file"].exists())
    processados = 0

    print(f"\n[2/3] Processando {total} fornecedor(es)...")

    for supplier, cfg in suppliers.items():
        rfi_path  = cfg["rfi_file"]
        anos_dest = cfg["anos"]   # { ano: (col_dest_ini, col_dest_fim) }

        if not rfi_path.exists():
            print(f"\n  [AVISO] {rfi_path.name} não encontrado — pulando {supplier}.")
            continue

        # Filtra anos que têm colunas de origem configuradas
        anos_validos = {ano: col_origem[ano] for ano in anos_dest if ano in col_origem}
        anos_sem_col = [ano for ano in anos_dest if ano not in col_origem]
        if anos_sem_col:
            print(f"\n  [AVISO] {supplier}: anos sem colunas de origem configuradas "
                  f"(aba Colunas): {anos_sem_col} — ignorados.")

        if not anos_validos:
            print(f"\n  [AVISO] {supplier}: nenhum ano válido para processar — pulando.")
            continue

        processados += 1
        print(f"\n  ({processados}/{total}) {supplier} — {rfi_path.name}")
        print(f"  Anos: {sorted(anos_validos)}")
        print(f"  {'─' * 55}")

        total_gravadas = 0

        for aba_ori, regras_aba in sorted(regras_por_aba.items()):
            print(f"\n    [{aba_ori}]")
            try:
                dados_por_dest = read_rfi_aba(rfi_path, aba_ori, anos_validos, regras_aba)
            except ValueError as e:
                print(f"    [ERRO] {e} — pulando esta aba.")
                continue

            for aba_dest, dados in dados_por_dest.items():
                # Exibe resumo das gravações
                n_copiar = sum(
                    1 for r in regras_aba if r.acao == "Copiar" and r.aba_destino == aba_dest
                    for _ in r.pares
                )
                n_somar  = sum(
                    1 for r in regras_aba if r.acao == "Somar"  and r.aba_destino == aba_dest
                )
                for ano, (cd_ini, cd_fim) in sorted(anos_dest.items()):
                    if ano in anos_validos:
                        print(f"      → {aba_dest} | {ano}: "
                              f"dest {get_column_letter(cd_ini)}–{get_column_letter(cd_fim)} | "
                              f"Copiar: {n_copiar} linhas, Somar: {n_somar} grupos")

                gravadas = write_to_painel(wb_painel, aba_dest, dados, anos_dest)
                total_gravadas += gravadas

        print(f"\n    Total de linhas gravadas no Painel: {total_gravadas}")

    # ── Salvar ───────────────────────────────────────────────────────────
    print(f"\n[3/3] Salvando: {output_file.name}...")
    wb_painel.save(output_file)
    print(f"      Arquivo salvo.")

    print(f"\n{'═' * 60}")
    print(f"  Concluído! {processados} fornecedor(es) processado(s).")
    print(f"  Saída: {output_file}")
    print(f"{'═' * 60}")


if __name__ == "__main__":
    main()
